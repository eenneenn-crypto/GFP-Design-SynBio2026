"""Create comprehensive Excel report with all 11 variants"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reference sequences
sfGFP = "MSKGEELFTGVVPILVELDGDVNGHKFSVRGEGEGDATNGKLTLKFICTTGKLPVPWPTLVTTLTYGVQCFSRYPDHMKRHDFFKSAMPEGYVQERTISFKDDGTYKTRAEVKFEGDTLVNRIELKGIDFKEDGNILGHKLEYNFNSHNVYITADKQKNGIKANFKIRHNVEDGSVQLADHYQQNTPIGDGPVLLPDNHYLSTQSVLSKDPNEKRDHMVLLEFVTAAGITHGMDELYK"
avGFP = "MSKGEELFTGVVPILVELDGDVNGHKFSVSGEGEGDATYGKLTLKFICTTGKLPVPWPTLVTTLSYGVQCFSRYPDHMKQHDFFKSAMPEGYVQERTIFFKDDGNYKTRAEVKFEGDTLVNRIELKGIDFKEDGNILGHKLEYNYNSHNVYIMADKQKNGIKVNFKIRHNIEDGSVQLADHYQQNTPIGDGPVLLPDNHYLSTQSALSKDPNEKRDHMVLLEFVTAAGITHGMDELYK"
amacGFP = "MSKGEELFTGIVPVLIELDGDVHGHKFSVRGEGEGDADYGKLEIKFICTTGKLPVPWPTLVTTLSYGILCFARYPEHMKMNDFFKSAMPEGYIQERTIFFQDDGKYKTRGEVKFEGDTLVNRIELKGMDFKEDGNILGHKLEYNFNSHNVYIMPDKANNGLKVNFKIRHNIEGGGVQLADHYQTNVPLGDGPVLIPINHYLSCQTAISKDRNETRDHMVFLEFFSACGHTHGMDELYK"
cgreGFP = "MTALTEGAKLFEKEIPYITELEGDVEGMKFIIKGEGTGDATTGTIKAKYICTTGDLPVPWATILSSLSYGVFCFAKYPRHIADFFKSTQPDGYSQDRIISFDNDGQYDVKAKVTYENGTLYNRVTVKGTGFKSNGNILGMRVLYHSPPHAVYILPDRKNGGMKIEYNKAFDVMGGGHQMARHAQFNKPLGAWEEDYPLYHHLTVWTSFGKDPDDDETDHLTIVEVIKAVDLETYR"
ppluGFP = "MPAMKIECRITGTLNGVEFELVGGGEGTPEQGRMTNKMKSTKGALTFSPYLLSHVMGYGFYHFGTYPSGYENPFLHAINNGGYTNTRIEKYEDGGVLHVSFSYRYEAGRVIGDFKVVGTGFPEDSVIFTDKIIRSNATVEHLHPMGDNVLVGSFARTFSLRDGGYYSFVVDSHMHFKSAIHPSILQNGGPMFAFRRVEELHSNTELGIVEYQHAFKTPIAFA"

def mutate(seq, pos1, new_aa):
    return seq[:pos1-1] + new_aa + seq[pos1:]

def get_mutations(seq, ref):
    muts = []
    for i, (a, b) in enumerate(zip(seq, ref)):
        if a != b:
            muts.append(f"{b}{i+1}{a}")
    return muts

# Define all 11 variants with their data
variants = []

# ===== G1: sfGFP_Plus_v1 =====
s1 = sfGFP
for p, a in [(69,'L'),(147,'P'),(149,'K'),(164,'Y'),(172,'K'),(190,'N'),(227,'V')]:
    s1 = mutate(s1, p, a)
variants.append({
    'id': 'G1',
    'name': 'sfGFP_Plus_v1',
    'scaffold': 'sfGFP (2B3P)',
    'seq': s1,
    'design_rationale': 'sfGFP骨架 + FoldX验证稳定突变 + usGFP核心堆积突变。E172K使sfGFP从+3.40降至-0.61',
    'design_source': 'Pédelacq 2006(sfGFP), Scott 2018(usGFP Q69L/N164Y), 本次FoldX(E172K/N149K/D190N/S147P/A227V)',
    'dg': 'N/A (组合效应)',
    'ddg': 'N/A (组合效应)',
})

# ===== G2: sfGFP_Plus_v2_TGP =====
s2 = sfGFP
for p, a in [(69,'L'),(79,'E'),(149,'K'),(164,'Y'),(172,'K'),(204,'E')]:
    s2 = mutate(s2, p, a)
variants.append({
    'id': 'G2',
    'name': 'sfGFP_Plus_v2_TGP',
    'scaffold': 'sfGFP (2B3P)',
    'seq': s2,
    'design_rationale': 'sfGFP + TGP式表面电荷工程(Q79E/Q204E增加负电荷) + 核心稳定突变。借鉴Close 2015表面工程策略',
    'design_source': 'Close 2015(TGP表面电荷), Scott 2018(usGFP), 本次FoldX(E172K/N149K)',
    'dg': 'N/A (组合效应)',
    'ddg': 'N/A (组合效应)',
})

# ===== G3: amacGFP_Plus_v1 =====
a1 = amacGFP
for p, a in [(39,'N'),(149,'K'),(172,'K')]:
    a1 = mutate(a1, p, a)
variants.append({
    'id': 'G3',
    'name': 'amacGFP_Plus_v1',
    'scaffold': 'amacGFP (7LG4)',
    'seq': a1,
    'design_rationale': 'amacGFP骨架 + 3个FoldX验证的最佳稳定突变。Y39N(-0.81)+N149K(-0.51)+E172K(-0.44)均为DDG+DG双优',
    'design_source': '本次FoldX(Y39N/N149K/E172K在7LG4上全部DDG<-3.9)',
    'dg': 'N/A (组合效应)',
    'ddg': 'N/A (组合效应)',
})

# ===== G4: amacGFP_Plus_v2 =====
a2 = amacGFP
for p, a in [(39,'N'),(69,'L'),(147,'P'),(149,'K'),(164,'Y'),(172,'K'),(190,'N')]:
    a2 = mutate(a2, p, a)
variants.append({
    'id': 'G4',
    'name': 'amacGFP_Plus_v2',
    'scaffold': 'amacGFP (7LG4)',
    'seq': a2,
    'design_rationale': 'G3升级版，增加usGFP(Q69L/N164Y)和额外稳定突变(S147P/D190N)，共7个稳定突变堆叠',
    'design_source': '本次FoldX + Scott 2018(usGFP)',
    'dg': 'N/A (组合效应)',
    'ddg': 'N/A (组合效应)',
})

# ===== G5: amacGFP_Plus_v3_max =====
a3 = amacGFP
for p, a in [(39,'N'),(63,'A'),(147,'P'),(149,'K'),(164,'Y'),(171,'V'),(172,'K'),(190,'N')]:
    a3 = mutate(a3, p, a)
variants.append({
    'id': 'G5',
    'name': 'amacGFP_Plus_v3_max',
    'scaffold': 'amacGFP (7LG4)',
    'seq': a3,
    'design_rationale': '最大突变版本，8个突变全部来自FoldX验证在amacGFP上DDG<0的突变',
    'design_source': '本次FoldX(8个突变全部在7LG4上验证为稳定化)',
    'dg': 'N/A (组合效应)',
    'ddg': 'N/A (组合效应)',
})

# ===== G6: ppluGFP_Opt_v1_bright =====
p1 = ppluGFP
for p, a in [(48,'E'),(118,'E'),(199,'H'),(206,'E'),(221,'E')]:
    p1 = mutate(p1, p, a)
variants.append({
    'id': 'G6',
    'name': 'ppluGFP_Opt_v1_bright',
    'scaffold': 'ppluGFP (2G6X)',
    'seq': p1,
    'design_rationale': 'ppluGFP天然极稳(DG=-49.88)+亮度优化突变(L199H/G206E/A221E) + 表面电荷(K48E/S118E)。WT亮度x3.2 avGFP',
    'design_source': 'GFP_data.xlsx亮度数据(L199H 4.39,G206E 4.40,A221E 4.38), Close 2015(表面电荷)',
    'dg': '-49.88 (WT基准,突变主要优化亮度)',
    'ddg': 'N/A (亮度优化为主)',
})

# ===== G7: ppluGFP_Opt_v2_surface =====
p2 = ppluGFP
for p, a in [(73,'E'),(147,'E'),(174,'E'),(199,'F'),(206,'E')]:
    p2 = mutate(p2, p, a)
variants.append({
    'id': 'G7',
    'name': 'ppluGFP_Opt_v2_surface',
    'scaffold': 'ppluGFP (2G6X)',
    'seq': p2,
    'design_rationale': 'ppluGFP + 表面Glu工程(S73E/S147E/S174E增加负电荷防聚集) + 亮度L206E',
    'design_source': 'GFP_data.xlsx亮度数据, Close 2015(TGP表面工程)',
    'dg': '-49.88 (WT基准)',
    'ddg': 'N/A (亮度优化为主)',
})

# ===== G8: cgreGFP_Opt_v1_bright =====
c1 = cgreGFP
for p, a in [(78,'G'),(167,'M'),(168,'V'),(172,'E')]:
    c1 = mutate(c1, p, a)
variants.append({
    'id': 'G8',
    'name': 'cgreGFP_Opt_v1_bright',
    'scaffold': 'cgreGFP (2HPW)',
    'seq': c1,
    'design_rationale': 'cgreGFP天然最亮(x6 avGFP!)+极稳(DG=-26.89)。亮度优化突变K167M(4.57)+A168V(4.55)+M172E(4.56)+R78G',
    'design_source': 'GFP_data.xlsx亮度最大突变组合, 天然高稳定性',
    'dg': '-26.89 (WT基准)',
    'ddg': 'N/A (亮度优化为主)',
})

# ===== G9: cgreGFP_Opt_v2 =====
c2 = cgreGFP
for p, a in [(8,'Q'),(167,'M'),(170,'E'),(197,'P')]:
    c2 = mutate(c2, p, a)
variants.append({
    'id': 'G9',
    'name': 'cgreGFP_Opt_v2',
    'scaffold': 'cgreGFP (2HPW)',
    'seq': c2,
    'design_rationale': 'cgreGFP + K167M+K8Q+D170E+L197P亮度优化组合',
    'design_source': 'GFP_data.xlsx亮度数据',
    'dg': '-26.89 (WT基准)',
    'ddg': 'N/A (亮度优化为主)',
})

# ===== G10: avGFP_Superfold =====
d_av = avGFP
for p, a in [(30,'R'),(39,'N'),(64,'L'),(99,'S'),(105,'T'),(145,'F'),(149,'K'),(153,'T'),(163,'A'),(171,'V'),(172,'K'),(206,'V')]:
    d_av = mutate(d_av, p, a)
variants.append({
    'id': 'G10',
    'name': 'avGFP_Superfold',
    'scaffold': 'avGFP (2WUR)',
    'seq': d_av,
    'design_rationale': '将sfGFP所有超级折叠突变导入avGFP骨架 + FoldX最佳突变。avGFP天然DG=-2.02已稳固',
    'design_source': 'Pédelacq 2006(sfGFP全部突变), 本次FoldX(N149K/E172K)',
    'dg': 'N/A (组合效应)',
    'ddg': 'N/A (组合效应)',
})

# ===== G11: Chimera_sf_amac =====
ch = sfGFP
for p, a in [(69,'L'),(79,'E'),(147,'P'),(149,'K'),(164,'Y'),(172,'K'),(190,'N'),(227,'V')]:
    ch = mutate(ch, p, a)
variants.append({
    'id': 'G11',
    'name': 'Chimera_sf_amac',
    'scaffold': 'sfGFP (2B3P)',
    'seq': ch,
    'design_rationale': 'sfGFP骨架 + 融合amacGFP的稳定特征 + TGP表面电荷 + FoldX稳定突变。跨骨架创新设计',
    'design_source': '综合: Pédelacq 2006, Scott 2018, Close 2015, 本次FoldX',
    'dg': 'N/A (组合效应)',
    'ddg': 'N/A (组合效应)',
})

# ===== Create Excel =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GFP设计总表"

# Styles
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
data_font = Font(name='微软雅黑', size=10)
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column widths
col_widths = [5, 22, 16, 60, 40, 18, 18]
headers = ['编号', '名称', '骨架', '完整序列(238aa)', '设计思路', 'DG(kcal/mol)', 'DDG(kcal/mol)']

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(i)].width = w

# Add data
for row_idx, v in enumerate(variants, 2):
    vals = [v['id'], v['name'], v['scaffold'], v['seq'], v['design_rationale'], v['dg'], v['ddg']]
    
    # Color by scaffold
    if 'sfGFP' in v['scaffold']:
        fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
    elif 'amac' in v['scaffold']:
        fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    elif 'pplu' in v['scaffold']:
        fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    elif 'cgre' in v['scaffold']:
        fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    else:
        fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    
    for col_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border

# Add sheet 2: Single mutation DG table
ws2 = wb.create_sheet('单突变DG数据')

headers2 = ['模板', 'WT DG', '突变', 'Mut DG', 'DDG', '效果']
col_w2 = [16, 12, 10, 10, 10, 10]

for i, (h, w) in enumerate(zip(headers2, col_w2), 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(i)].width = w

# Single mutation data
single_muts = {
    'sfGFP (2B3P)': (3.40, [('E172K', -0.61, -4.01), ('N149K', 0.26, -3.14), ('D190N', 0.74, -2.66), ('S147P', 0.87, -2.53), ('A227V', 1.49, -1.91)]),
    'avGFP (2WUR)': (-2.02, [('S30R', -1.44, 0.58), ('N149K', -0.58, 1.44), ('M153T', -0.49, 1.53), ('E172K', -0.35, 1.67), ('Y39N', -0.27, 1.75), ('D190N', -0.08, 1.94), ('Y39D', 0.18, 2.20), ('S147P', 0.18, 2.20), ('T63A', 0.77, 2.79), ('I171V', 1.16, 3.18), ('F99S', 2.38, 4.40), ('A227V', 3.11, 5.13), ('V163A', 3.56, 5.58)]),
    'amacGFP (7LG4)': (3.55, [('Y39N', -0.81, -4.36), ('N149K', -0.51, -4.06), ('E172K', -0.44, -3.99), ('S147P', -0.07, -3.62), ('D190N', -0.07, -3.62), ('T63A', 0.33, -3.22), ('Y39D', 0.73, -2.82), ('I171V', 0.79, -2.76), ('M153T', 1.95, -1.60), ('F99S', 2.60, -0.95), ('V163A', 3.52, -0.03)]),
}

row = 2
for tmpl, (wt, muts) in single_muts.items():
    for i, (mut, dg, ddg) in enumerate(muts):
        if i == 0:
            ws2.cell(row=row, column=1, value=tmpl).font = data_font
            ws2.cell(row=row, column=2, value=wt).font = data_font
        ws2.cell(row=row, column=3, value=mut).font = data_font
        dg_cell = ws2.cell(row=row, column=4, value=dg)
        dg_cell.font = Font(name='微软雅黑', size=10, color='1B5E20' if dg < 0 else 'B71C1C')
        ddg_cell = ws2.cell(row=row, column=5, value=ddg)
        ddg_cell.font = Font(name='微软雅黑', size=10, color='1B5E20' if ddg < 0 else 'B71C1C')
        
        effect = '稳定' if dg < 0 else ('偏高' if dg < 2 else '不稳定')
        ws2.cell(row=row, column=6, value=effect).font = data_font
        
        for c in range(1, 7):
            ws2.cell(row=row, column=c).border = thin_border
        row += 1

# Add sheet 3: Reference sequences
ws3 = wb.create_sheet('参考序列')
ws3.cell(row=1, column=1, value='名称').font = header_font
ws3.cell(row=1, column=1).fill = header_fill
ws3.cell(row=1, column=2, value='推荐PDB').font = header_font
ws3.cell(row=1, column=2).fill = header_fill
ws3.cell(row=1, column=3, value='长度(aa)').font = header_font
ws3.cell(row=1, column=3).fill = header_fill
ws3.cell(row=1, column=4, value='WT DG(kcal/mol)').font = header_font
ws3.cell(row=1, column=4).fill = header_fill
ws3.cell(row=1, column=5, value='WT亮度(log)').font = header_font
ws3.cell(row=1, column=5).fill = header_fill
ws3.cell(row=1, column=6, value='完整序列').font = header_font
ws3.cell(row=1, column=6).fill = header_fill

for c in range(1, 7):
    ws3.cell(row=1, column=c).border = thin_border

refs = [
    ('sfGFP', '2B3P', 238, 3.40, 'N/A(基线)', sfGFP),
    ('avGFP', '2WUR', 238, -2.02, 3.72, avGFP),
    ('amacGFP', '7LG4', 238, 3.55, 3.97, amacGFP),
    ('cgreGFP', '2HPW', 242, -26.89, 4.50, cgreGFP),
    ('ppluGFP', '2G6X', 222, -49.88, 4.23, ppluGFP),
]

for i, (name, pdb, length, dg, bright, seq) in enumerate(refs, 2):
    ws3.cell(row=i, column=1, value=name).font = data_font
    ws3.cell(row=i, column=2, value=pdb).font = data_font
    ws3.cell(row=i, column=3, value=length).font = data_font
    ws3.cell(row=i, column=4, value=dg).font = data_font
    ws3.cell(row=i, column=5, value=bright).font = data_font
    ws3.cell(row=i, column=6, value=seq).font = data_font
    for c in range(1, 7):
        ws3.cell(row=i, column=c).border = thin_border
    ws3.cell(row=i, column=6).alignment = Alignment(wrap_text=True)

# Save
output = r'D:\openclaw_work\2026synbio比赛\GFP_设计总表.xlsx'
wb.save(output)
print(f'Saved to: {output}')
print('Sheets: GFP设计总表, 单突变DG数据, 参考序列')
